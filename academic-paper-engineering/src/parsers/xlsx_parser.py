"""XLSX 解析器 - 将 Excel 电子表格解析为 Table IR"""

from pathlib import Path
from typing import Dict, List, Optional


class XlsxParser:
    """解析 .xlsx 格式的 Excel 文件，提取表格数据"""

    def __init__(self):
        self.ir = {
            "source_file": "",
            "sheets": []
        }

    def parse(self, file_path: str) -> Dict:
        """解析 XLSX 文件，返回 Table IR"""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        self.ir["source_file"] = path.name

        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_data = self._parse_sheet(ws, sheet_name)
                self.ir["sheets"].append(sheet_data)

        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        return self.ir

    def _parse_sheet(self, ws, sheet_name: str) -> Dict:
        """解析单个工作表"""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {
                "sheet_name": sheet_name,
                "table_id": "",
                "data_range": "",
                "headers": [],
                "rows": [],
                "summary_rows": [],
                "notes": [],
                "suggested_caption": ""
            }

        # 识别表头行（第一行非空）
        header_row = rows[0]
        headers = []
        for col_idx, value in enumerate(header_row):
            if value is not None:
                headers.append({
                    "text": str(value).strip(),
                    "colspan": 1,
                    "rowspan": 1
                })

        # 提取数据行
        data_rows = []
        summary_keywords = ["平均", "标准差", "总计", "合计", "均值", "average",
                           "std", "total", "sum", "mean"]
        summary_rows = []

        for row_idx, row in enumerate(rows[1:], start=1):
            if all(v is None for v in row):
                continue

            cells = []
            is_summary = False
            for value in row:
                if value is None:
                    cells.append({"text": "", "type": "empty"})
                else:
                    text = str(value).strip()
                    cell_type = "text"
                    if isinstance(value, (int, float)):
                        cell_type = "number"
                    cells.append({"text": text, "type": cell_type})

                    # 检查是否为汇总行
                    if any(kw in text.lower() for kw in summary_keywords):
                        is_summary = True

            row_data = {"row": row_idx, "cells": cells}
            if is_summary:
                summary_rows.append(row_data)
            else:
                data_rows.append(row_data)

        return {
            "sheet_name": sheet_name,
            "table_id": f"table_{sheet_name}",
            "data_range": f"A1:{chr(64 + len(headers))}{len(rows)}",
            "headers": [{"row": 0, "cells": headers}],
            "rows": data_rows,
            "summary_rows": summary_rows,
            "notes": [],
            "suggested_caption": f"{sheet_name} 数据表"
        }

    def to_table_ir(self, sheet_data: Dict, table_number: int = 1) -> Dict:
        """将工作表数据转换为标准 Table IR 格式"""
        return {
            "id": f"table_{table_number:03d}",
            "number": table_number,
            "caption": sheet_data.get("suggested_caption", ""),
            "label": f"tab:{sheet_data['sheet_name']}",
            "placement": "htbp",
            "headers": sheet_data.get("headers", []),
            "rows": sheet_data.get("rows", []),
            "footnotes": [],
            "notes": "\n".join(sheet_data.get("notes", [])),
            "column_count": len(sheet_data.get("headers", [{}])[0].get("cells", [])),
            "row_count": len(sheet_data.get("rows", [])),
            "type": "simple"
        }
